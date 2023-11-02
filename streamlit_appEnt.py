# ! Las dependencia, rutas y codigos que se usan en la terminal de anaconda

# cd OneDrive - Grupo Bancolombia\Workspace\FicsAppStreamLit\
# cd Workspace\FIC StreamLit
# streamlit run streamlit_appEnt.py

# pip install -r requirements.txt


# pip install    qgrid     -i https://artifactory.apps.bancolombia.com/api/pypi/python-org/simple --trusted-host artifactory.apps.bancolombia.com

# pip install -r requirements.txt -i https://artifactory.apps.bancolombia.com/api/pypi/python-org/simple --trusted-host artifactory.apps.bancolombia.com



# ! Los Dataframe con terminacion "NoDupl" es para la visualizacion NO USAR en el excel final

import pandas as pd
import streamlit as st
# import plotly.express as px
from openpyxl import Workbook as Wb
from io import BytesIO

from xlsxwriter import Workbook
from tempfile import NamedTemporaryFile

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# from openpyxl.writer.excel import save_virtual_workbook

from io import StringIO


from PIL import Image
# from pyxlsb import open_workbook as open_xlsb


from openpyxl.styles import Font

import asyncio



from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)




st.set_page_config(
    page_title="FICs App",
    page_icon="img/LogoBancolombiaNegro.png",
    # layout="wide",
    initial_sidebar_state="expanded",

)


customized_button = st.markdown("""
    <style >
    # .stDownloadButton, div.stButton {text-align:center}
    .stDownloadButton button, div.stButton > button:first-child {
        background-color: #ff4b4b;
        color:#ffffff;
        padding-left: 20px;
        padding-right: 20px;
        transition: opacity 0.5s ease-in-out;
    }

    .css-1n543e5:focus:not(:active) {
    border-color: rgb(255, 75, 75);
    color: #ffffff;
    }

    .stDownloadButton button:hover, div.stButton > button:hover {
        font-size: 2.5rem;
        background-color: #ffffff;
        color: #ff4b4b;
        border-color: #ff4b4b;
    }
    .stDownloadButton button:focus:not(:active) {
        background-color: #ffffff;
        color: #ff4b4b;
        border-color: #ff4b4b;
    }
    .stDownloadButton button:visited {
        background-color: #ff4b4b;
        color: #ffffff;
        border-color: #ff4b4b;
    }
        }
    </style>""", unsafe_allow_html=True)



empty_left, contents, empty_right = st.columns([0.5, 3, 0.5])

with contents:
    st.header("Reporte de competencia industria ㅤㅤㅤㅤ local de fondos")


empty_left, contents, empty_right = st.columns([1.9, 3, 0.1])

with contents:
    st.markdown("Fecha Corte: 06 30 2023")


st.text(" ")


img = Image.open("img/investment3.jpeg")
st.image(img, use_column_width=True)


excel_file = "MODELO.xlsb"
sheet_name = "BD"

df = pd.read_excel(excel_file,
                   sheet_name= sheet_name,
                   header=0,
                   usecols = "A:AF",
                   )


dfTiposFondos = pd.read_excel("BD ASSET CLASS.xlsx",
                           sheet_name= "Hoja1",
                           header= 0)

dfIndustriaLocal = pd.read_excel( "BDIndustriaLocalFICs.xlsx",
                   sheet_name= "BD 30Abr2023",
                   header=1,
                   usecols = "A:Z",
                   )




dictNombresCortos = dict(zip(dfIndustriaLocal['concatenar'],
                                  dfIndustriaLocal['Nombre Corto']
                                  ))

dictComisionAdmin = dict(zip(dfIndustriaLocal['concatenar'],
                                  dfIndustriaLocal['Comisión admin(%)']
                                  ))

dictEntidadCorto = dict(zip(dfIndustriaLocal['Nombre Entidad'],
                                  dfIndustriaLocal['Nombre Corto Entidad']
                                  ))



df.columns = df.columns.str.replace('peer_group.Tipo de participación ficha técnica', 'Tipo de participación ficha técnica')
df.columns = df.columns.str.replace('Cons. id Part.', 'ID Participacion')
df.columns = df.columns.str.replace('fichas.Dur_Años', 'Duracion Años')

df = df.assign(Nombre_Fondo_Corto= "" )
df = df.assign(Nombre_Entidad_Corto= "" )
df = df.assign(RB_mensual = "" )
df = df.assign(RB_semestral = "" )
df = df.assign(RB_Ytd = "" )
df = df.assign(RB_1Y = "" )
df = df.assign(RB_3Y = "" )
df = df.assign(RB_5Y = "" )
df = df.assign(Select= False )


print("Corriendo Nombre Corto Fondo")
for i in range(df.shape[0]):

    nombreFondo = df["Llave"][i]
    if nombreFondo in dictNombresCortos:
        nombreCorto = dictNombresCortos[nombreFondo]
        df.at[i, "Nombre_Fondo_Corto"] = nombreCorto
    else:
        df.at[i, "Nombre_Fondo_Corto"] = nombreFondo
        print(nombreFondo)


print("Corriendo Comision")
for i in range(df.shape[0]):

    nombreFondo = df["Llave"][i]
    if nombreFondo in dictComisionAdmin:
        comisionAdmin = dictComisionAdmin[nombreFondo]
        df.at[i, "Comision_Admin"] = comisionAdmin
    else:
        df.at[i, "Comision_Admin"] = "-"



print("Corriendo Nombre Corto Entidad")
for i in range(df.shape[0]):

    nombreEntidad = df["Nombre Entidad"][i]
    if nombreEntidad in dictEntidadCorto:
        nombreCorto = dictEntidadCorto[nombreEntidad]
        df.at[i, "Nombre_Entidad_Corto"] = nombreCorto

    else:
        df.at[i, "Nombre_Entidad_Corto"] = nombreEntidad





print("Corriendo Rentabilidades brutas")
def calcularRB(rentabilidad, comision):

    try:
        rentB = (1+rentabilidad)/(1+(comision/100))-1

    except:
        rentB = "ND"

    return rentB


for i in range(df.shape[0]):

    nombreFondo = df["Llave"][i]
    if nombreFondo in dictComisionAdmin :

        comision = dictComisionAdmin[nombreFondo]

        df.at[i, "RB_mensual"] = calcularRB(df["RN.mensual"][i], comision)
        df.at[i, "RB_semestral"] = calcularRB(df["RN.semestral"][i], comision)
        df.at[i, "RB_Ytd"] = calcularRB(df["RN.Ytd"][i], comision)
        df.at[i, "RB_1Y"] = calcularRB(df["RN. 1Y"][i], comision)
        df.at[i, "RB_3Y"] = calcularRB(df["RN. 3Y"][i], comision)
        df.at[i, "RB_5Y"] = calcularRB(df["RN. 5Y"][i], comision)

    else:
        df.at[i, "RB_mensual"] = "-"
        df.at[i, "RB_semestral"] = "-"
        df.at[i, "RB_Ytd"] = "-"
        df.at[i, "RB_1Y"] = "-"
        df.at[i, "RB_3Y"] = "-"
        df.at[i, "RB_5Y"] = "-"





dfTiposFondosNoDupl = dfTiposFondos.drop_duplicates(subset=["NOMBRE NEGOCIO"], keep='first')


# ! Emparejar fondo con su tipo:

df = df.assign(ASSET_CLASS= "" )

nombreFondos70 = df["Nombre Negocio"].unique().tolist()

nombresDBTiposFondos = dfTiposFondos["NOMBRE NEGOCIO"].unique().tolist()


listaTiposFondos = dfTiposFondos["ASSET CLASS"].unique().tolist()


diccionarioTiposFondos = dict(zip(dfTiposFondosNoDupl['NOMBRE NEGOCIO'],
                                  dfTiposFondosNoDupl['ASSET CLASS']
                                  ))


# diccionarioEntidadCorta = dict(zip(dfTiposFondosNoDupl['NOMBRE NEGOCIO'],
#                                   dfTiposFondosNoDupl['NOMBRE CORTO ADMINISTRADORA']
#                                   ))

st.text(" ")
st.text(" ")
st.text(" ")



col1, col2, col3 = st.columns([1, 2, 0.1])

with col2:

    st.subheader("¿Cómo usar los filtros?")

empty_left, contents, empty_right = st.columns([2.6, 3, 0.1])

with contents:
    st.markdown("(Tutorial)")

st.text(" ")


with st.expander("Tipos de filtrado"):
    
    st.subheader("1. Filtrar solo por nombre")
    imgf1 = Image.open("img/Inkedfiltersnombre_LI.jpg")
    st.image(imgf1, use_column_width=True)

    st.text(" ")
    st.subheader("2. Filtrar solo por asset class")
    imgf2 = Image.open("img/Inkedfiltersasset_LI.jpg")
    st.image(imgf2, use_column_width=True)

    st.text(" ")
    st.subheader("3. Filtrar por asset class y por nombre")
    imgf3 = Image.open("img/Inkedfiltersnombreasset_LI.jpg")
    st.image(imgf3, use_column_width=True)

    st.text(" ")
    st.subheader("4. No filtrar por por orden nombre-asset")
    imgf3 = Image.open("img/Inkedfiltersnosense.png")
    st.image(imgf3, use_column_width=True)
    

    st.text(" ")



with st.expander("Abreviaturas"):

    empty_left, contents, empty_right = st.columns([2.75, 2.2, 2])



    with empty_left:


        st.markdown("SN - SENTENCIAS NACION")
        st.markdown("PP - PACTO DE PERMANENCIA")

    with contents:
        st.markdown("RF - RENTA FIJA")
        st.markdown("LP - LARGO PLAZO")

    with empty_right:
        st.markdown("TS - TASA FIJA")
        st.markdown("COL - COLOMBIA")

st.success(
"""
   ㅤAl final los fondos que queden dentro de recuadro **\"Nombre Negocio\"** seran los descargados
""" )



for i in range(df.shape[0]):


    nombreFondo = df["Nombre Negocio"][i]

    if nombreFondo in diccionarioTiposFondos:

        tipoFondo= diccionarioTiposFondos[nombreFondo]
        df.at[i, 'ASSET_CLASS'] = tipoFondo
    else:
        df.at[i, 'ASSET_CLASS'] = "INDEFINIDO"



st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")

empty_left, contents, empty_right = st.columns([0.95, 3, 0.75])

with contents:
    st.subheader("Descargue nuestros :red[_fondos sugeridos_]")

st.text(" ")






# ! Descargar por Excel
@st.cache_data(ttl=3600)

def to_excel(df, numeroFondos):
    # output = BytesIO()
    # writer = pd.ExcelWriter(output, engine='xlsxwriter')
    # df.to_excel(writer, index=False, sheet_name='Sheet1')
    # workbook = writer.book
    # worksheet = writer.sheets['Sheet1']
    # format1 = workbook.add_format({'num_format': '0.00'})
    # worksheet.set_column('A:A', None)
    # writer.close()
    # processed_data = output.getvalue()
    # return processed_data





    wb = load_workbook('template.xlsx') 
    ws = wb.active

    rows = dataframe_to_rows(df)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
             ws.cell(row=r_idx+10, column=c_idx, value=value)




    ws.delete_cols(1)
    ws.delete_rows(12)


    if numeroFondos == "70":
        headers = ws["A11":"AH11"]
    elif numeroFondos == "All":
        headers = ws["A11":"AL11"]


    fuente = Font( bold=True , color='FFFFFF')

    for row in headers:
        for cell in row:
            cell.font = fuente
            # cell.style.alignment.horizontal = 'center'
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")

    output = BytesIO()
    wb.save(output)
    # workbook = Wb()

    # with NamedTemporaryFile() as tmp:
    #     workbook.save(tmp.name)

    data = output.getvalue()
    return data



df_filtrado = df[["ASSET_CLASS", "Nombre_Entidad_Corto", "Nombre_Fondo_Corto",	"Valor fondo",
         "# Inversionistas", 	"Tipo de participación ficha técnica", "Comision_Admin",	"Duracion Años",	"Tipo de participación (TP)",
         "RN.mensual",	"RN.semestral",	"RN.Ytd",	"RN. 1Y", "RN. 3Y", "RN. 5Y",
         "RB_mensual", "RB_semestral", "RB_Ytd", "RB_1Y", "RB_3Y", "RB_5Y" ,"V.mensual",
         "V.semestral", "V.Ytd", "V. 1Y", "V. 3Y",	"V. 5Y", "Sharpe.1Y", "Sharpe.3Y",
         "Sharpe.5Y",	"# veces con RN<0 semana",	"# veces con RN<0 mes",
         "# veces con RN<0 YtD",	"# veces con RN<0 1Y"
]]


# df_filtrado.columns = df_filtrado.columns.str.replace('Nombre_Corto', 'Nombre Fondo')


col1, col2, col3 = st.columns([1.20, 2, 0.1])



df_filtrado = df_filtrado.sort_values(by="Nombre_Fondo_Corto", ascending=True)


with col2:
    # with open(excel_file, 'rb') as my_file:
    st.download_button(label='Generar Informe Sugeridos',
                       data=to_excel(df_filtrado, "70") ,
                       file_name= 'FondosSugeridos.xlsx'
                       )


st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")


empty_left, contents, empty_right = st.columns([0.65, 3, 0.1])

with contents:
    st.subheader("Filtre y seleccione los que usted desee 🔍")

empty_left, contents, empty_right = st.columns([1.75, 3, 0.1])

with contents:
    st.markdown("(Base de fondos sugeridos)")


st.text(" ")

dfCheckColumn = pd.DataFrame

# async def guardaChecks(df): #Simple async def
#     df = pd.DataFrame
#     dfCheckColumn = df["Select"]
#     return dfCheckColumn

# async def main_def(df):
    
#     df = st.data_editor(
#         filter_dataframe(df),
#         column_order=("Select","Nombre_Entidad_Corto", "Nombre_Fondo_Corto", "ASSET_CLASS"),
#         column_config={
#             "Select": st.column_config.CheckboxColumn(
#                 help="Selecciona tus **fondos**",
#                 default=False,
#             )
#         },
#         disabled=["widgets"],
#         hide_index=True,
#     )
#     await asyncio.gather(guardaChecks(df))
    


# asyncio.run(main_def())
# # The function you wait for must include async
# # The function you use await must include async
# # The function you use await must run by asyncio.run(THE_FUNC())







def filter_dataframe(df: pd.DataFrame) -> pd.DataFrame:

    modify = st.checkbox("Add filters")

    if not modify:
        return df

    df = df.copy()
    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    dfColumns = ["ASSET_CLASS", "Nombre Negocio"]
    with modification_container:


        to_filter_columns = st.multiselect("Filtrar por: ", dfColumns)
        for column in to_filter_columns:

            left, right = st.columns((1, 20))
            left.write("↳")
            # Treat columns with < 10 unique values as categorical

            # if is_categorical_dtype(df[column]) or df[column].nunique() < 10:

            user_cat_input = right.multiselect(
                    f"{column}",
                    df[column].unique(),
                    default=(df[column].to_list())[0],

                )


            df = df[df[column].isin(user_cat_input)]
            # elif is_numeric_dtype(df[column]):
            #     _min = float(df[column].min())
            #     _max = float(df[column].max())
            #     step = (_max - _min) / 100
            #     user_num_input = right.slider(
            #         f"Values for {column}",
            #         _min,
            #         _max,
            #         (_min, _max),
            #         step=step,
            #     )
            #     df = df[df[column].between(*user_num_input)]
            # elif is_datetime64_any_dtype(df[column]):
            #     user_date_input = right.date_input(
            #         f"Values for {column}",
            #         value=(
            #             df[column].min(),
            #             df[column].max(),
            #         ),
            #     )
            #     if len(user_date_input) == 2:
            #         user_date_input = tuple(map(pd.to_datetime, user_date_input))
            #         start_date, end_date = user_date_input
            #         df = df.loc[df[column].between(start_date, end_date)]
            # else:
            #     user_text_input = right.text_input(
            #         f"Substring or regex in {column}",
            #     )
            #     if user_text_input:
            #         df = df[df[column].str.contains(user_text_input)]





    return df



async def seleccionarFondo(df):
    
    df = st.data_editor(
        df,
        column_order=("Select","Nombre_Entidad_Corto", "Nombre_Fondo_Corto", "ASSET_CLASS"),
        column_config={
            "Select": st.column_config.CheckboxColumn(
                help="Selecciona tus **fondos**",
                default=False,
            )
        },
        disabled=["widgets"],
        hide_index=True,
    )

    return df


dictSelect = {"a": 1}


async def crearDictSelect(df, llave, filter_dataframe):

    dfPermanecer = seleccionarFondo(df)

    dfFunctionFilter = filter_dataframe(dfPermanecer)

    df_mask = dfPermanecer['Select']==True

    filtered_df = dfPermanecer[df_mask] 
    

    dictPermanecer = dict(zip( filtered_df[llave],
                       filtered_df['Select']
                      ))



    global dictSelect

    dictSelect.update(dictPermanecer)

    
    df = df.reset_index()
    dfSelect = df

    for i in range(dfSelect.shape[0]):

        nombreFondo = dfSelect[llave][i]

        if nombreFondo in dictPermanecer:

            dfSelect.at[i, "Select"] = True

    dictTrueSelect = {}
    for fondo in dictPermanecer:
        if dictPermanecer[fondo] == True:
            dictTrueSelect.update({fondo: dictPermanecer[fondo]})

            

    return dfSelect, dictPermanecer, dictTrueSelect

# dfNoDupl= df.drop_duplicates(subset=["Nombre Negocio"], keep='first')


# dfSelect, dictSelect70,dictTrueSelect70 = crearDictSelect(dfNoDupl, 'Nombre_Fondo_Corto', filter_dataframe)
# dfNoDupl = dfSelect






# df70Vacio = pd.DataFrame()

# for fondo in dictTrueSelect70:
    
#     df70 = df.loc[df["Nombre_Fondo_Corto"]== fondo]
#     df70Vacio = pd.concat([df70Vacio, df70], axis=0)


# col1, col2, col3 = st.columns(3)


# with col1:

#     st.download_button(label='Generar Informe 70 Check',
#                        data=to_excel(df70Vacio) ,
#                        file_name= 'Informe70Check.xlsx'
#                        )


df_downl =filter_dataframe(df)

df_downlNoDupl = df_downl.drop_duplicates(subset=["Nombre_Fondo_Corto"], keep='first')








st.dataframe(df_downlNoDupl[['Nombre_Entidad_Corto','Nombre_Fondo_Corto'
                             , "ASSET_CLASS"
                            ]],  hide_index=True )


df_downl = df_downl[["ASSET_CLASS", "Nombre_Entidad_Corto", "Nombre_Fondo_Corto",	"Valor fondo",
         "# Inversionistas", 	"Tipo de participación ficha técnica", "Comision_Admin",	"Duracion Años",	"Tipo de participación (TP)",
         "RN.mensual",	"RN.semestral",	"RN.Ytd",	"RN. 1Y", "RN. 3Y", "RN. 5Y",
         "RB_mensual", "RB_semestral", "RB_Ytd", "RB_1Y", "RB_3Y", "RB_5Y" ,"V.mensual",
         "V.semestral", "V.Ytd", "V. 1Y", "V. 3Y",	"V. 5Y", "Sharpe.1Y", "Sharpe.3Y",
         "Sharpe.5Y",	"# veces con RN<0 semana",	"# veces con RN<0 mes",
         "# veces con RN<0 YtD",	"# veces con RN<0 1Y"
]]







col1, col2, col3 = st.columns(3)



df_downl2023 = df_downl.sort_values(by="Nombre_Fondo_Corto", ascending=True)


with col1:
    st.download_button(label='Generar Informe',
                                    data=to_excel(df_downl, "70") ,
                                    file_name= 'MisFondos.xlsx')




st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")


# ! SIF 2023!!!

def filter_dataframeSIF(df: pd.DataFrame) -> pd.DataFrame:

    modifySIF = st.checkbox("Add filters SIF")

    if not modifySIF:
        return df

    df = df.copy()
    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass

        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    dfColumns2023 = ["ASSET_CLASS", "Nombre_Fondo_Corto"]
    with modification_container:


        to_filter_columns2023 = st.multiselect("Filtra por: ", dfColumns2023,key="SIF")
        for column in to_filter_columns2023:

            left, right = st.columns((1, 20))
            left.write("↳")

            user_cat_input = right.multiselect(
                    f"{column}",
                    df[column].unique(),
                    default=(df[column].to_list())[0],

                )


            df = df[df[column].isin(user_cat_input)]


    return df

st.text(" ")
st.text(" ")
st.text(" ")

empty_left, contents, empty_right = st.columns([0.6, 2, 0.1])

with contents:
    st.subheader("_Base total industria local de fondos_")

empty_left, contents, empty_right = st.columns([1.5, 3, 0.1])

with contents:
    st.markdown("(Fuente: Reporte 523 Superfinanciera)")


sheetSIF2023 = "SIF_2023Actualizado"
excelSIF2023 = sheetSIF2023 + ".xlsx"

#   Original:                       "SIF_2023Actualizado"
#   Sin "Concatenar Duplicado":     "SIF_2023NoDuplAct"

@st.cache_data(ttl=3600)
def load_data(excel,sheet):
    # Carga tu DataFrame aquí
    df = pd.read_excel(excel,
                sheet_name= sheet,
                  header= 0)

    return df



dfSIF2023 = load_data(excelSIF2023,sheetSIF2023)


dfSIF2023 = dfSIF2023.assign(Select= False )


dfSIF2023NoDupl = dfSIF2023.drop_duplicates(subset=["Nombre_Fondo_Corto"], keep='first')

# dfSIF2023NoDupl, dictSelectTodos, dictTrueSelectTodos = crearDictSelect(
#     dfSIF2023NoDupl, "Nombre_Fondo_Corto", filter_dataframeSIF)


# dfTodosVacio = pd.DataFrame()



# for fondo in dictTrueSelectTodos:
    
#     dfTodos = dfSIF2023.loc[dfSIF2023["Nombre_Fondo_Corto"]== fondo]
#     dfTodosVacio = pd.concat([dfTodosVacio, dfTodos], axis=0)



# col1, col2, col3 = st.columns(3)




# col1, col2, col3 = st.columns([2, 1, 0.1])

# with col1:
#     st.download_button(label='Generar Informe Todos Check',
#                        data=to_excel(dfTodosVacio) ,
#                        file_name= 'SIFInformeCheck.xlsx'
#                        )



df_downl2023 =filter_dataframeSIF(dfSIF2023)

df_downl2023 = df_downl2023[["concatenar",	"Fecha corte",	"ASSET_CLASS",	"Nombre_Entidad_Corto",	"Nombre_Fondo_Corto",	"ID Participacion",	"Núm. unidades",	"Valor unidad para las operaciones del día t",	"Valor fondo al cierre del día t",	"Núm. Invers.",	"Comision_Admin",	"Rentab. dia",	"Rentab. mes",	"Rentab. sem",	"Rentab. Ultaño",	"Rentab_Ytd",	"Rentab_3Y",	"Rentab_5Y",	"RB_mensual",	"RB_semestral",	"RB_Ytd",	"RB_1Y",	"RB_3Y",	"RB_5Y",	"V_mensual",	"V_semestral",	"V_Ytd",	"V_1Y",	"V_3Y",	"V_5Y",	"Sharpe_1Y",	"Sharpe_3Y",	"Sharpe_5Y",	"Rentab_Neg_semana",	"Rentab_Neg_mes",	"Rentab_Neg_YtD",	"Rentab_Neg_Semestre",	"Rentab_Neg_1Y"

]]

df_downl2023NoDupl = df_downl2023.drop_duplicates(subset=["Nombre_Fondo_Corto"], keep='first')




st.dataframe(df_downl2023NoDupl[['Nombre_Entidad_Corto','Nombre_Fondo_Corto',
                                 "ASSET_CLASS"
                                ]],  hide_index=True )





col1, col2, col3 = st.columns(3)


df_downl2023 = df_downl2023.sort_values(by="Nombre_Fondo_Corto", ascending=True)



with col1:

    st.download_button(label='Generar Informe SIF',
                       data=to_excel(df_downl2023, "All") ,
                       file_name= 'SIFInforme.xlsx'
                       )


# def load_data(sheets_url):
#     csv_url = sheets_url.replace("/edit#gid=", "/export?format=csv&gid=")
#     return pd.read_csv(csv_url)


# df = load_data(st.secrets["public_gsheets_url"])

# for row in df.itertuples():
#     st.write(f"{row.name} has a :{row.pet}:")

# @st.cache_data(ttl=3600)
# def load_data(url):
#     return  pd.read_csv(url, dtype=str).fillna("")



# Connect to the Google Sheet
# sheet_id = "1tfWAudn1Hkd3TizWbeif7ZdJHEQYH8UpWQv18q7gJxw"
# sheet_name = "1816189210"
# url = f"<https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}>"
# df = load_data(url)

st.text(" ")
st.text(" ")
 

st.text(" ")
st.text(" ")

st.info(
    """
    ㅤㅤㅤㅤㅤㅤㅤㅤㅤ¿Inquietudes? ㅤ [Dejanos saber a través de este link!](https://forms.gle/7oGTStkjUXNeztLV9)
    """,
    icon="👀",
)
# st.write(df)