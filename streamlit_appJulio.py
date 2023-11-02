# ! Las dependencia, rutas y codigos que se usan en la terminal de anaconda

# cd OneDrive - Grupo Bancolombia\Workspace\FicsAppStreamLit\
# streamlit run streamlit_app.py

# pip install -r requirements.txt


# pip install    qgrid     -i https://artifactory.apps.bancolombia.com/api/pypi/python-org/simple --trusted-host artifactory.apps.bancolombia.com

# pip install -r requirements.txt -i https://artifactory.apps.bancolombia.com/api/pypi/python-org/simple --trusted-host artifactory.apps.bancolombia.com



# ! Los Dataframe con terminacion "NoDupl" es para la visualizacion NO USAR en el excel final

import pandas as pd
import streamlit as st
# import plotly.express as px
from openpyxl import Workbook as Wb
from io import BytesIO

from xlsxwriter import Workbook
from tempfile import NamedTemporaryFile

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# from openpyxl.writer.excel import save_virtual_workbook

from io import StringIO


from PIL import Image
# from pyxlsb import open_workbook as open_xlsb


from openpyxl.styles import Font

import asyncio



from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)




st.set_page_config(
    page_title="FICs App",
    page_icon="img/LogoBancolombiaNegro.png",
    # layout="wide",
    initial_sidebar_state="expanded",

)

hide_github_icon = """
#GithubIcon {
  visibility: hidden;
}
"""

# st.markdown(hide_github_icon, unsafe_allow_html=True)



customized_button = st.markdown("""
    <style >
    # .stDownloadButton, div.stButton {text-align:center}
    .stDownloadButton button, div.stButton > button:first-child {
        background-color: #ff4b4b;
        color:#ffffff;
        padding-left: 20px;
        padding-right: 20px;
        transition: opacity 0.5s ease-in-out;
    }

    .css-1n543e5:focus:not(:active) {
    border-color: rgb(255, 75, 75);
    color: #ffffff;
    }

    .stDownloadButton button:hover, div.stButton > button:hover {
        font-size: 2.5rem;
        background-color: #ffffff;
        color: #ff4b4b;
        border-color: #ff4b4b;
    }
    .stDownloadButton button:focus:not(:active) {
        background-color: #ffffff;
        color: #ff4b4b;
        border-color: #ff4b4b;
    }
    .stDownloadButton button:visited {
        background-color: #ff4b4b;
        color: #ffffff;
        border-color: #ff4b4b;
    }
        }
    </style>""", unsafe_allow_html=True)



empty_left, contents, empty_right = st.columns([1, 3, 0.5])

with contents:
    st.header("Reporte de competencia industria local de fondos")


empty_left, contents, empty_right = st.columns([1.9, 3, 0.1])

with contents:
    st.markdown("Fecha Corte: 07 31 2023")


st.text(" ")


img = Image.open("img/investment3.jpeg")
st.image(img, use_column_width=True)


excel_file = "Informe de competencia FICs 31072023 Todos los fondos.xlsx"
sheet_name = "Informe Completo"


dfSIF = pd.read_excel(excel_file,
                   sheet_name= sheet_name,
                   header=0,
                   usecols = "A:AI",
                   )


dfSIF["Fecha corte"] = "31/07/2023"


colsACambiar = ["Valor fondo", 
                "Comisión",	
                "Duración",	
                "RN.mensual",	
                "RN.semestral",	
                "RN.Ytd", 
                "RN. 1Y",  
                "RN. 3Y", 	
                "RN. 5Y",	
                "RB.mensual",	
                "RB.semestral",	
                "RB.Ytd",	
                "RB. 1Y",	
                "RB. 3Y",	
                "RB. 5Y",	
                "V.mensual",	
                "V.semestral",	
                "V.Ytd",	
                "V. 1Y",	
                "V. 3Y",	
                "V. 5Y"]


dfSIF['Valor fondo'] =  dfSIF['Valor fondo'].map("{:,.2f}".format)
dfSIF['Comisión'] =  dfSIF['Comisión'].map("{:,.2f}".format)
dfSIF['Duración'] =  dfSIF['Duración'].map("{:,.2f}".format)
dfSIF['RN.mensual'] =  dfSIF['RN.mensual'].map("{:,.2f}".format)
dfSIF['RN.semestral'] =  dfSIF['RN.semestral'].map("{:,.2f}".format)
dfSIF['RN.Ytd'] =  dfSIF['RN.Ytd'].map("{:,.2f}".format)
dfSIF['RN. 1Y'] =  dfSIF['RN. 1Y'].map("{:,.2f}".format)
dfSIF['RN. 3Y'] =  dfSIF['RN. 3Y'].map("{:,.2f}".format)
dfSIF['RN. 5Y'] =  dfSIF['RN. 5Y'].map("{:,.2f}".format)
dfSIF['RB.mensual'] =  dfSIF['RB.mensual'].map("{:,.2f}".format)
dfSIF['RB.semestral'] =  dfSIF['RB.semestral'].map("{:,.2f}".format)
dfSIF['RB.Ytd'] =  dfSIF['RB.Ytd'].map("{:,.2f}".format)
dfSIF['RB. 1Y'] =  dfSIF['RB. 1Y'].map("{:,.2f}".format)
dfSIF['RB. 3Y'] =  dfSIF['RB. 3Y'].map("{:,.2f}".format)
dfSIF['RB. 5Y'] =  dfSIF['RB. 5Y'].map("{:,.2f}".format)
dfSIF['V.mensual'] =  dfSIF['V.mensual'].map("{:,.2f}".format)
dfSIF['V.semestral'] =  dfSIF['V.semestral'].map("{:,.2f}".format)
dfSIF['V.Ytd'] =  dfSIF['V.Ytd'].map("{:,.2f}".format)
dfSIF['V. 1Y'] =  dfSIF['V. 1Y'].map("{:,.2f}".format)
dfSIF['V. 3Y'] =  dfSIF['V. 3Y'].map("{:,.2f}".format)
dfSIF['V. 5Y'] =  dfSIF['V. 5Y'].map("{:,.2f}".format)


dfSIF.replace({"nan": "ND"})



filtered_df = dfSIF.dropna()

st.text(" ")
st.text(" ")



col1, col2, col3 = st.columns([1.5, 2, 0.1])

with col2:

    st.subheader("Tutorial")


with st.expander("Hacer busquedas"):
    st.subheader("Usted puede buscar en cada tabla")
    st.markdown("1. De **_click encima_** de la tabla.")
    st.markdown("2. Use **_Ctrl + F_** para abrir buscador")
    st.markdown("3. Ingrese las palabras clave que quiere buscar")



    imgf1 = Image.open("img/busqueda.png")
    st.image(imgf1, use_column_width=True)


with st.expander("Tipos de filtrado"):
    
    st.subheader("1. Filtrar solo por nombre")
    imgf1 = Image.open("img/Inkedfiltersnombre_LI.jpg")
    st.image(imgf1, use_column_width=True)

    st.text(" ")
    st.subheader("2. Filtrar solo por asset class")
    imgf2 = Image.open("img/Inkedfiltersasset_LI.jpg")
    st.image(imgf2, use_column_width=True)

    st.text(" ")
    st.subheader("3. Filtrar por asset class y por nombre")
    imgf3 = Image.open("img/Inkedfiltersnombreasset_LI.jpg")
    st.image(imgf3, use_column_width=True)

    st.text(" ")
    st.subheader("4. No filtrar por por orden nombre-asset")
    imgf3 = Image.open("img/Inkedfiltersnosense.png")
    st.image(imgf3, use_column_width=True)
    



with st.expander("Abreviaturas"):

    empty_left, contents, empty_right = st.columns([2.75, 2.2, 2])



    with empty_left:


        st.markdown("SN - SENTENCIAS NACION")
        st.markdown("PP - PACTO DE PERMANENCIA")

    with contents:
        st.markdown("RF - RENTA FIJA")
        st.markdown("LP - LARGO PLAZO")

    with empty_right:
        st.markdown("TS - TASA FIJA")
        st.markdown("COL - COLOMBIA")


st.success(
"""
   ㅤAl final los fondos que queden dentro de recuadro **\"Nombre Negocio\"** seran los descargados
""" )

st.success(
"""
   ㅤLa duración solo está disponible para la muestra de fondos sugeridos
""" )




st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")



empty_left, contents, empty_right = st.columns([0.95, 3, 0.75])

with contents:
    st.subheader("Descargue nuestros :red[_fondos sugeridos_]")

st.text(" ")






# ! Descargar por Excel
@st.cache_data(ttl=3600)

def to_excel(df, numeroFondos):
    # output = BytesIO()
    # writer = pd.ExcelWriter(output, engine='xlsxwriter')
    # df.to_excel(writer, index=False, sheet_name='Sheet1')
    # workbook = writer.book
    # worksheet = writer.sheets['Sheet1']
    # format1 = workbook.add_format({'num_format': '0.00'})
    # worksheet.set_column('A:A', None)
    # writer.close()
    # processed_data = output.getvalue()
    # return processed_data


    wb = load_workbook('template.xlsx') 
    ws = wb.active

    rows = dataframe_to_rows(df)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
             ws.cell(row=r_idx, column=c_idx, value=value)


 

    ws.delete_cols(1)
    ws.delete_rows(2)


    output = BytesIO()
    wb.save(output)
    # workbook = Wb()

    # with NamedTemporaryFile() as tmp:
    #     workbook.save(tmp.name)

    data = output.getvalue()
    return data



col1, col2, col3 = st.columns([1.20, 2, 0.1])


with col2:
    st.download_button(label='Generar Informe Sugeridos',
                       data=to_excel(filtered_df, "70") ,
                       file_name= 'FondosSugeridos.xlsx'
                       )


st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")


empty_left, contents, empty_right = st.columns([0.70, 3, 0.1])

with contents:
    st.subheader("Filtre y seleccione de la base de fondos ㅤㅤㅤㅤㅤㅤㅤㅤ sugeridos 🔍")

empty_left, contents, empty_right = st.columns([2, 3, 0.1])

with contents:
    st.markdown("(Base de fondos sugeridos)")


st.text(" ")






def filter_dataframe(df: pd.DataFrame) -> pd.DataFrame:

    modify = st.checkbox("Add filters")

    if not modify:
        return df

    df = df.copy()
    # Try to convert datetimes into a standard format (datetime, no timezone)
#    for col in df.columns:
#        if is_object_dtype(df[col]):
#            try:
#                df[col] = pd.to_datetime(df[col])
#            except Exception:
#                pass

#        if is_datetime64_any_dtype(df[col]):
#            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    dfColumns = ["Asset Class", "NOMBRE CORTO FONDO"]
    with modification_container:


        to_filter_columns = st.multiselect("Filtrar por: ", dfColumns)
        for column in to_filter_columns:

            left, right = st.columns((1, 20))
            left.write("↳")
            # Treat columns with < 10 unique values as categorical

            # if is_categorical_dtype(df[column]) or df[column].nunique() < 10:

            user_cat_input = right.multiselect(
                    f"{column}",
                    df[column].unique(),
                    default=(df[column].to_list())[0],

                )


            df = df[df[column].isin(user_cat_input)]


    return df



df_downl =filter_dataframe(filtered_df)


st.dataframe(df_downl[['NOMBRE CORTO ADMINISTRADORA','NOMBRE CORTO FONDO',
                                  "Asset Class"
                                ]],  hide_index=True )


col1, col2, col3 = st.columns(3)



with col1:
    st.download_button(label='Generar Informe',
                                    data=to_excel(df_downl, "70") ,
                                    file_name= 'MisFondos.xlsx')


st.text(" ")
st.text(" ")
st.text(" ")


# ! SIF 2023!!!

def filter_dataframeSIF(df: pd.DataFrame) -> pd.DataFrame:

    modifySIF = st.checkbox("Add filters SIF")

    if not modifySIF:
        return df

    df = df.copy()
    # Try to convert datetimes into a standard format (datetime, no timezone)
#    for col in df.columns:
#        if is_object_dtype(df[col]):
#            try:
#                df[col] = pd.to_datetime(df[col])
#            except Exception:
#                pass
#
#        if is_datetime64_any_dtype(df[col]):
#            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()

    dfColumns2023 = ["Asset Class", "NOMBRE CORTO FONDO"]
    with modification_container:


        to_filter_columns2023 = st.multiselect("Filtra por: ", dfColumns2023,key="SIF")
        for column in to_filter_columns2023:

            left, right = st.columns((1, 20))
            left.write("↳")

            user_cat_input = right.multiselect(
                    f"{column}",
                    df[column].unique(),
                    default=(df[column].to_list())[0],

                )


            df = df[df[column].isin(user_cat_input)]


    return df

st.text(" ")
st.text(" ")
st.text(" ")

empty_left, contents, empty_right = st.columns([0.6, 2, 0.1])

with contents:
    st.subheader("_Base total industria local de fondos_")

empty_left, contents, empty_right = st.columns([1.5, 3, 0.1])

with contents:
    st.markdown("(Fuente: Reporte 523 Superfinanciera)")


sheetSIF2023 = "MODELO Rutas Julian.xlsb"
excelSIF2023 = "Informe Completo"

#   Original:                       "SIF_2023Actualizado"
#   Sin "Concatenar Duplicado":     "SIF_2023NoDuplAct"






dfdownlSIF =filter_dataframeSIF(dfSIF)


dfdownlSIFNoDupl = dfdownlSIF.drop_duplicates(subset=["NOMBRE CORTO FONDO"], keep='first')



st.dataframe(dfdownlSIFNoDupl[['NOMBRE CORTO ADMINISTRADORA','NOMBRE CORTO FONDO',
                                  "Asset Class"
                                ]],  hide_index=True )





col1, col2, col3 = st.columns(3)





with col1:

    st.download_button(label='Generar Informe SIF',
                       data=to_excel(dfdownlSIF, "All") ,
                       file_name= 'SIFInforme.xlsx'
                       )




st.text(" ")
st.text(" ")
st.text(" ")
st.text(" ")


st.info(
    """
    ㅤㅤㅤㅤㅤSi presenta alguna inquietud al respecto, puede escribirnos al correo:ㅤㅤㅤㅤㅤㅤㅤㅤㅤㅤㅤㅤㅤ 
    Gerencia_Desarrollo_Negocio_AM@bancolombia.com.co
    """,
    icon="👀",
)
